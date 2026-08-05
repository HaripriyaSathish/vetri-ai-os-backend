from rest_framework import viewsets, permissions
from rest_framework.views import APIView
from rest_framework.response import Response
from django.contrib.auth import get_user_model
from core.email_utils import send_email
from django.utils import timezone
from django.db import models
import json

from .models import (
    Batch, LessonPlan, DailySchedule, Attendance, Assignment,
    MockInterviewQuestion, AssignmentSubmission, Report, MockInterviewSession, Message,
)
from .serializers import (
    BatchSerializer, LessonPlanSerializer, DailyScheduleSerializer, AttendanceSerializer,
    AssignmentSerializer, MockInterviewQuestionSerializer, AssignmentSubmissionSerializer,
    ReportSerializer, MockInterviewSessionSerializer, MessageSerializer,
)
from rest_framework import serializers
import secrets
from django.http import HttpResponseRedirect
from django.shortcuts import get_object_or_404
from .models import ClassRecording, RecordingView, Enrollment
from .models import AbsenceNotification
from datetime import date, timedelta
from .report_excel import build_zone_report_excel
from rest_framework.exceptions import PermissionDenied
from .models import Holiday
from .serializers import HolidaySerializer

User = get_user_model()


def display_name(user):
    full = f"{user.first_name} {user.last_name}".strip()
    return full or user.username


class BatchViewSet(viewsets.ModelViewSet):
    serializer_class = BatchSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.role in ('admin', 'management'):
            return Batch.objects.all()
        return Batch.objects.filter(trainer=user)

    def perform_create(self, serializer):
        serializer.save(trainer=self.request.user)

    def perform_destroy(self, instance):
        if self.request.user.role != 'management':
            raise PermissionDenied("Only business team can remove batches.")
        instance.delete()

class BatchStudentsView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, batch_id):
        student_ids = (
            Attendance.objects.filter(batch_id=batch_id)
            .order_by('student_id')
            .values_list('student_id', flat=True)
            .distinct()
        )
        results = []
        for sid in student_ids:
            student = User.objects.get(id=sid)
            records = Attendance.objects.filter(batch_id=batch_id, student_id=sid)
            total = records.count()
            present = records.filter(status='present').count()
            results.append({
                'id': student.id,
                'username': student.username,
                'email': student.email,
                'official_email': student.official_email,
                'first_name': student.first_name,
                'last_name': student.last_name,
                'attendance_percentage': round((present / total) * 100, 1) if total > 0 else None,
                'days_recorded': total,
            })
        return Response(results)
class BatchLessonTimelineView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, batch_id):
        plans = LessonPlan.objects.filter(batch_id=batch_id).order_by('date')
        data = [{'date': p.date, 'topic': p.topic} for p in plans]
        return Response(data)


def get_batch_eligibility(batch_id):
    student_ids = (
        Attendance.objects.filter(batch_id=batch_id)
        .order_by('student_id')
        .values_list('student_id', flat=True)
        .distinct()
    )
    total_assignments = Assignment.objects.filter(batch_id=batch_id).count()
    results = []

    for sid in student_ids:
        student = User.objects.get(id=sid)
        att_records = Attendance.objects.filter(batch_id=batch_id, student_id=sid)
        total_days = att_records.count()
        present_days = att_records.filter(status='present').count()
        attendance_pct = round((present_days / total_days) * 100, 1) if total_days > 0 else 0

        submissions = AssignmentSubmission.objects.filter(student_id=sid, assignment__batch_id=batch_id, verified=True)
        on_time_count = sum(1 for s in submissions if s.submitted_at.date() <= s.assignment.due_date)
        all_on_time = (
            total_assignments > 0
            and submissions.count() >= total_assignments
            and on_time_count >= total_assignments
        )

        eligible = attendance_pct >= 85 and all_on_time

        session = MockInterviewSession.objects.filter(batch_id=batch_id, student_id=sid).first()

        results.append({
    'student_id': sid,
    'username': student.username,
    'first_name': student.first_name,
    'last_name': student.last_name,
    'email': student.email,
    'attendance_percentage': attendance_pct,
    'assignments_submitted': submissions.count(),
    'total_assignments': total_assignments,
    'all_on_time': all_on_time,
    'eligible': eligible,
    'invited': session is not None,
    'session_id': session.id if session else None,
    'scheduled_datetime': session.scheduled_datetime if session else None,
    'meeting_link': session.meeting_link if session else None,
    'attended': session.attended if session else None,
    'score': session.score if session else None,
    'feedback': session.feedback if session else None,
})
    return results

class BatchTrainingLogView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, batch_id):
        start = request.query_params.get('start')
        end = request.query_params.get('end')
        if not start or not end:
            return Response({"detail": "start and end dates are required."}, status=400)

        from datetime import datetime as dt
        start_date = dt.strptime(start, '%Y-%m-%d').date()
        end_date = dt.strptime(end, '%Y-%m-%d').date()

        batch = Batch.objects.get(id=batch_id)
        student_ids = (
            Attendance.objects.filter(batch_id=batch_id)
            .order_by('student_id')
            .values_list('student_id', flat=True)
            .distinct()
        )
        students = [User.objects.get(id=sid) for sid in student_ids]

        holidays = {h.date: h.reason for h in Holiday.objects.filter(batch_id=batch_id, date__gte=start_date, date__lte=end_date)}
        schedules = DailySchedule.objects.filter(batch_id=batch_id, date__gte=start_date, date__lte=end_date)
        lesson_plans = {p.date: p.topic for p in LessonPlan.objects.filter(batch_id=batch_id, date__gte=start_date, date__lte=end_date)}

        rows = []
        sno = 1
        current = start_date
        while current <= end_date:
            # Sunday is always off — never appears in the report at all
            if current.weekday() == 6:
                current += timedelta(days=1)
                continue

            # Holiday — one explicit row noting it, no per-student rows
            if current in holidays:
                rows.append({
                    'sno': sno,
                    'date': str(current),
                    'trainer_name': display_name(batch.trainer),
                    'trainee_name': '—',
                    'status': f"Holiday — {holidays[current]}",
                    'training_mode': batch.training_mode,
                    'course_name': batch.course_name or batch.name,
                    'timings': '—',
                    'programming_language': batch.programming_language or '—',
                    'topics_covered': '—',
                })
                sno += 1
                current += timedelta(days=1)
                continue

            # A regular day only appears if attendance was actually marked that day —
            # no attendance marked means no class happened, so it's silently skipped
            attendance_on_date = {
                a.student_id: a.status
                for a in Attendance.objects.filter(batch_id=batch_id, date=current)
            }
            if not attendance_on_date:
                current += timedelta(days=1)
                continue

            schedule = schedules.filter(date=current).first()
            timing = f"{schedule.start_time.strftime('%I:%M %p')} - {schedule.end_time.strftime('%I:%M %p')}" if schedule else "—"
            topic = lesson_plans.get(current, '—')

            for student in students:
                status = attendance_on_date.get(student.id, 'Not Marked')
                rows.append({
                    'sno': sno,
                    'date': str(current),
                    'trainer_name': display_name(batch.trainer),
                    'trainee_name': display_name(student),
                    'status': status.capitalize() if status != 'Not Marked' else status,
                    'training_mode': batch.training_mode,
                    'course_name': batch.course_name or batch.name,
                    'timings': timing,
                    'programming_language': batch.programming_language or '—',
                    'topics_covered': topic,
                })
                sno += 1

            current += timedelta(days=1)

        return Response(rows)

class MockEligibilityView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        batch_id = request.query_params.get('batch_id')
        if not batch_id:
            return Response({"detail": "batch_id is required."}, status=400)
        return Response(get_batch_eligibility(batch_id))


class SendMockInviteView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        batch_id = request.data.get('batch_id')
        interview_date = request.data.get('interview_date')
        window_start = request.data.get('window_start', '10:00 AM')
        window_end = request.data.get('window_end', '5:00 PM')
        subject_template = (request.data.get('subject') or '').strip()
        body_template = (request.data.get('body') or '').strip()
        cc_raw = request.data.get('cc', '')

        if not batch_id:
            return Response({"detail": "batch_id is required."}, status=400)
        if not interview_date:
            return Response({"detail": "interview_date is required."}, status=400)

        cc_list = [c.strip() for c in cc_raw.split(',') if c.strip()] if cc_raw else None

        batch = Batch.objects.get(id=batch_id)
        eligibility = get_batch_eligibility(batch_id)
        invited = []

        default_subject = f"Mock Interview - You're Eligible! - {batch.name}"
        default_body = (
            "<p>Hi {{full_name}},</p>"
            f"<p>Congratulations! You've completed all assignments on time and maintained "
            "{{attendance_percentage}}% attendance in <strong>" + batch.name + "</strong>. "
            "You're eligible for the Mock Interview round.</p>"
            f"<p>Please be ready and available on <strong>{interview_date}</strong>, "
            f"between <strong>{window_start}</strong> and <strong>{window_end}</strong>.</p>"
            "<p>Your exact time slot and the Microsoft Teams meeting link will be shared with you individually shortly.</p>"
            "<p>Best of luck!</p>"
        )

        subject_to_use = subject_template or default_subject
        body_to_use = body_template or default_body

        for e in eligibility:
            if e['eligible'] and not e['invited']:
                MockInterviewSession.objects.create(batch=batch, student_id=e['student_id'])

                full_name = f"{e['first_name']} {e['last_name']}".strip() or e['username']
                personalized_subject = subject_to_use.replace('{{full_name}}', full_name)
                personalized_body = (
                    body_to_use
                    .replace('{{full_name}}', full_name)
                    .replace('{{attendance_percentage}}', str(e['attendance_percentage']))
                )

                sent = send_email(
                    to=e['email'],
                    subject=personalized_subject,
                    html_body=personalized_body,
                    cc=cc_list,
                )
                if sent:
                    invited.append(e['username'])

        return Response({"invited": invited, "count": len(invited)})

class ScheduleIndividualInterviewView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        session_id = request.data.get('session_id')
        scheduled_datetime = request.data.get('scheduled_datetime')
        meeting_link = request.data.get('meeting_link', '')

        if not session_id or not scheduled_datetime:
            return Response({"detail": "session_id and scheduled_datetime are required."}, status=400)

        try:
            session = MockInterviewSession.objects.get(id=session_id)
        except MockInterviewSession.DoesNotExist:
            return Response({"detail": "Session not found."}, status=404)

        session.scheduled_datetime = scheduled_datetime
        session.meeting_link = meeting_link
        session.save()

        student = session.student

        try:
            formatted_time = session.scheduled_datetime.strftime('%B %d, %Y at %I:%M %p')
        except Exception:
            formatted_time = str(session.scheduled_datetime)

        link_block = (
            f"<p><strong>Join via Microsoft Teams:</strong> <a href='{meeting_link}'>{meeting_link}</a></p>"
            if meeting_link else ""
        )

        send_email(
            to=student.email,
            subject=f"Your Mock Interview Time - {session.batch.name}",
            html_body=(
                f"<p>Hi {student.username},</p>"
                f"<p>Your mock interview for <strong>{session.batch.name}</strong> is confirmed.</p>"
                f"<p><strong>Date & Time:</strong> {formatted_time}</p>"
                f"{link_block}"
                f"<p>Please join a few minutes early. Best of luck!</p>"
            ),
        )

        return Response({"detail": "Scheduled and emailed."})

class MockInterviewSessionViewSet(viewsets.ModelViewSet):
    serializer_class = MockInterviewSessionSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = MockInterviewSession.objects.all()

    def get_queryset(self):
        batch_id = self.request.query_params.get('batch_id')
        qs = MockInterviewSession.objects.all()
        if batch_id:
            qs = qs.filter(batch_id=batch_id)
        return qs


class LessonPlanViewSet(viewsets.ModelViewSet):
    serializer_class = LessonPlanSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        batch_id = self.request.query_params.get('batch')
        qs = LessonPlan.objects.all()
        if batch_id:
            qs = qs.filter(batch_id=batch_id)
        return qs


class DailyScheduleViewSet(viewsets.ModelViewSet):
    serializer_class = DailyScheduleSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = DailySchedule.objects.all()

    def get_queryset(self):
        batch_id = self.request.query_params.get('batch_id')
        qs = DailySchedule.objects.all()
        if batch_id:
            qs = qs.filter(batch_id=batch_id)
        return qs


class AttendanceViewSet(viewsets.ModelViewSet):
    serializer_class = AttendanceSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = Attendance.objects.all()

    def get_queryset(self):
        qs = Attendance.objects.all()
        batch_id = self.request.query_params.get('batch_id')
        date = self.request.query_params.get('date')
        if batch_id:
            qs = qs.filter(batch_id=batch_id)
        if date:
            qs = qs.filter(date=date)
        return qs

    def perform_create(self, serializer):
        batch = serializer.validated_data['batch']
        student = serializer.validated_data['student']

        already_enrolled = Attendance.objects.filter(batch=batch, student=student).exists()

        if not already_enrolled:
            current_count = (
                Attendance.objects.filter(batch=batch)
                .order_by('student_id')
                .values_list('student_id', flat=True)
                .distinct()
                .count()
            )
            if current_count >= batch.max_students:
                raise serializers.ValidationError(
                    f"This batch has reached its maximum capacity of {batch.max_students} students."
                )

        serializer.save(marked_by=self.request.user)


class AssignmentViewSet(viewsets.ModelViewSet):
    serializer_class = AssignmentSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = Assignment.objects.all()

    def get_queryset(self):
        batch_id = self.request.query_params.get('batch_id')
        qs = Assignment.objects.all()
        if batch_id:
            qs = qs.filter(batch_id=batch_id)
        return qs


class MockInterviewQuestionViewSet(viewsets.ModelViewSet):
    serializer_class = MockInterviewQuestionSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = MockInterviewQuestion.objects.all()

    def get_queryset(self):
        batch_id = self.request.query_params.get('batch_id')
        qs = MockInterviewQuestion.objects.all()
        if batch_id:
            qs = qs.filter(batch_id=batch_id)
        return qs


class AssignmentSubmissionViewSet(viewsets.ModelViewSet):
    serializer_class = AssignmentSubmissionSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = AssignmentSubmission.objects.all()

    def get_queryset(self):
        qs = AssignmentSubmission.objects.all()
        student_id = self.request.query_params.get('student_id')
        assignment_id = self.request.query_params.get('assignment_id')
        batch_id = self.request.query_params.get('batch_id')
        if student_id:
            qs = qs.filter(student_id=student_id)
        if assignment_id:
            qs = qs.filter(assignment_id=assignment_id)
        if batch_id:
            qs = qs.filter(assignment__batch_id=batch_id)
        return qs

    def perform_update(self, serializer):
        if 'verified' in self.request.data and self.request.data['verified']:
            serializer.save(verified_at=timezone.now())
        else:
            serializer.save()

class ReportViewSet(viewsets.ModelViewSet):
    serializer_class = ReportSerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = Report.objects.all()

    def get_queryset(self):
        batch_id = self.request.query_params.get('batch_id')
        qs = Report.objects.all()
        if batch_id:
            qs = qs.filter(batch_id=batch_id)
        return qs


from calendar import monthrange
from datetime import date


class MonthlyAttendanceView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, batch_id):
        year = int(request.query_params.get('year'))
        month = int(request.query_params.get('month'))

        _, days_in_month = monthrange(year, month)
        start_date = date(year, month, 1)
        end_date = date(year, month, days_in_month)

        records = Attendance.objects.filter(
            batch_id=batch_id, date__gte=start_date, date__lte=end_date
        ).select_related('student')

        student_ids = (
            Attendance.objects.filter(batch_id=batch_id)
            .order_by('student_id')
            .values_list('student_id', flat=True)
            .distinct()
        )

        result = []
        for sid in student_ids:
            student = User.objects.get(id=sid)
            day_status = {}
            for day in range(1, days_in_month + 1):
                record = records.filter(student_id=sid, date=date(year, month, day)).first()
                day_status[str(day)] = record.status if record else None
            result.append({
                'student': display_name(student),
                'days': day_status,
            })

        return Response({
            'batch_id': batch_id,
            'year': year,
            'month': month,
            'days_in_month': days_in_month,
            'students': result,
        })    

class BatchZoneReportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, batch_id):
        start = request.query_params.get('start')
        end = request.query_params.get('end')

        if not start or not end:
            return Response({"detail": "start and end dates are required."}, status=400)

        batch = Batch.objects.get(id=batch_id)
        eligibility_data = get_batch_eligibility(batch_id)

        daily_tasks = Assignment.objects.filter(batch_id=batch_id, category='task')
        total_daily_tasks = daily_tasks.count()

        mini_projects = Assignment.objects.filter(batch_id=batch_id, category='mini_project')
        total_mini_projects = mini_projects.count()

        main_projects = Assignment.objects.filter(batch_id=batch_id, category='main_project')
        total_main_projects = main_projects.count()

        seminars = Assignment.objects.filter(batch_id=batch_id, category='seminar')
        total_seminars = seminars.count()

        schedule = DailySchedule.objects.filter(batch_id=batch_id).first()
        timing = f"{schedule.start_time.strftime('%I:%M %p')} - {schedule.end_time.strftime('%I:%M %p')}" if schedule else "—"

        rows = []
        sno = 1
        for e in eligibility_data:
            sid = e['student_id']

            att_records_period = Attendance.objects.filter(batch_id=batch_id, student_id=sid, date__gte=start, date__lte=end)
            total_days_period = att_records_period.count()
            present_days_period = att_records_period.filter(status='present').count()

            completed_daily_tasks = AssignmentSubmission.objects.filter(
                student_id=sid, assignment__batch_id=batch_id, assignment__category='task',
                submitted_at__date__gte=start, submitted_at__date__lte=end, verified=True,
            ).count()
            daily_task_pct = round((completed_daily_tasks / total_daily_tasks) * 100) if total_daily_tasks > 0 else 0

            completed_mini_projects = AssignmentSubmission.objects.filter(
                student_id=sid, assignment__batch_id=batch_id, assignment__category='mini_project',
                submitted_at__date__gte=start, submitted_at__date__lte=end, verified=True,
            ).count()
            mini_project_pct = round((completed_mini_projects / total_mini_projects) * 100) if total_mini_projects > 0 else 0

            completed_main_projects = AssignmentSubmission.objects.filter(
                student_id=sid, assignment__batch_id=batch_id, assignment__category='main_project',
                submitted_at__date__gte=start, submitted_at__date__lte=end, verified=True,
            ).count()
            main_project_pct = round((completed_main_projects / total_main_projects) * 100) if total_main_projects > 0 else 0

            completed_seminars = AssignmentSubmission.objects.filter(
                student_id=sid, assignment__batch_id=batch_id, assignment__category='seminar',
                submitted_at__date__gte=start, submitted_at__date__lte=end, verified=True,
            ).count()
            seminar_pct = round((completed_seminars / total_seminars) * 100) if total_seminars > 0 else 0

            zone = 'Safe Zone' if e['eligible'] else 'Danger Zone'

            trainee_full_name = f"{e['first_name']} {e['last_name']}".strip() or e['username']
            rows.append({
                'sno': sno,
                'trainer_name': display_name(batch.trainer),
                'trainee_name': trainee_full_name,
                'zone': zone,
                'batch': batch.course_name or batch.name,
                'timings': timing,
                'total_class_days': total_days_period,
                'total_present_days': present_days_period,
                'attendance_percentage': e['attendance_percentage'],
                'assigned_daily_tasks': total_daily_tasks,
                'completed_daily_tasks': completed_daily_tasks,
                'daily_task_percentage': daily_task_pct,
                'assigned_mini_projects': total_mini_projects,
                'completed_mini_projects': completed_mini_projects,
                'mini_project_percentage': mini_project_pct,
                'assigned_main_projects': total_main_projects,
                'completed_main_projects': completed_main_projects,
                'main_project_percentage': main_project_pct,
                'assigned_seminars': total_seminars,
                'completed_seminars': completed_seminars,
                'seminar_percentage': seminar_pct,
            })
            sno += 1

        return Response(rows)

class BatchFullZoneReportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, batch_id):
        batch = Batch.objects.get(id=batch_id)
        eligibility_data = get_batch_eligibility(batch_id)

        daily_tasks = Assignment.objects.filter(batch_id=batch_id, category='task')
        total_daily_tasks = daily_tasks.count()

        mini_projects = Assignment.objects.filter(batch_id=batch_id, category='mini_project')
        total_mini_projects = mini_projects.count()

        main_projects = Assignment.objects.filter(batch_id=batch_id, category='main_project')
        total_main_projects = main_projects.count()

        seminars = Assignment.objects.filter(batch_id=batch_id, category='seminar')
        total_seminars = seminars.count()

        schedule = DailySchedule.objects.filter(batch_id=batch_id).first()
        timing = f"{schedule.start_time.strftime('%I:%M %p')} - {schedule.end_time.strftime('%I:%M %p')}" if schedule else "—"

        rows = []
        sno = 1
        for e in eligibility_data:
            sid = e['student_id']

            att_records = Attendance.objects.filter(batch_id=batch_id, student_id=sid)
            total_days = att_records.count()
            present_days = att_records.filter(status='present').count()

            completed_daily_tasks = AssignmentSubmission.objects.filter(
                student_id=sid, assignment__batch_id=batch_id, assignment__category='task', verified=True,
            ).count()
            daily_task_pct = round((completed_daily_tasks / total_daily_tasks) * 100) if total_daily_tasks > 0 else 0

            completed_mini_projects = AssignmentSubmission.objects.filter(
                student_id=sid, assignment__batch_id=batch_id, assignment__category='mini_project', verified=True,
            ).count()
            mini_project_pct = round((completed_mini_projects / total_mini_projects) * 100) if total_mini_projects > 0 else 0

            completed_main_projects = AssignmentSubmission.objects.filter(
                student_id=sid, assignment__batch_id=batch_id, assignment__category='main_project', verified=True,
            ).count()
            main_project_pct = round((completed_main_projects / total_main_projects) * 100) if total_main_projects > 0 else 0

            completed_seminars = AssignmentSubmission.objects.filter(
                student_id=sid, assignment__batch_id=batch_id, assignment__category='seminar', verified=True,
            ).count()
            seminar_pct = round((completed_seminars / total_seminars) * 100) if total_seminars > 0 else 0

            zone = 'Safe Zone' if e['eligible'] else 'Danger Zone'

            trainee_full_name = f"{e['first_name']} {e['last_name']}".strip() or e['username']
            rows.append({
                'sno': sno,
                'trainer_name': display_name(batch.trainer),
                'trainee_name': trainee_full_name,
                'zone': zone,
                'batch': batch.course_name or batch.name,
                'timings': timing,
                'total_class_days': total_days,
                'total_present_days': present_days,
                'attendance_percentage': e['attendance_percentage'],
                'assigned_daily_tasks': total_daily_tasks,
                'completed_daily_tasks': completed_daily_tasks,
                'daily_task_percentage': daily_task_pct,
                'assigned_mini_projects': total_mini_projects,
                'completed_mini_projects': completed_mini_projects,
                'mini_project_percentage': mini_project_pct,
                'assigned_main_projects': total_main_projects,
                'completed_main_projects': completed_main_projects,
                'main_project_percentage': main_project_pct,
                'assigned_seminars': total_seminars,
                'completed_seminars': completed_seminars,
                'seminar_percentage': seminar_pct,
            })
            sno += 1

        return Response(rows)
class StudentProfileView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, batch_id, student_id):
        try:
            batch = Batch.objects.get(id=batch_id)
            student = User.objects.get(id=student_id, role='student')
        except (Batch.DoesNotExist, User.DoesNotExist):
            return Response({"detail": "Batch or student not found."}, status=404)

        # A trainer may only view students in their own batch; admin/management see any.
        if request.user.role == 'trainer' and batch.trainer_id != request.user.id:
            return Response({"detail": "Not authorized to view this batch."}, status=403)

        attendance_qs = Attendance.objects.filter(batch=batch, student=student)
        total_days = attendance_qs.count()
        present_days = attendance_qs.filter(status='present').count()
        attendance_pct = round((present_days / total_days) * 100, 1) if total_days > 0 else None

        category_breakdown = []
        for cat_key, cat_label in Assignment.CATEGORY_CHOICES:
            assignments = Assignment.objects.filter(batch=batch, category=cat_key).order_by('due_date')
            rows = []
            for a in assignments:
                sub = AssignmentSubmission.objects.filter(assignment=a, student=student).first()
                rows.append({
                    'id': a.id,
                    'title': a.title,
                    'due_date': a.due_date,
                    'submitted': bool(sub),
                    'submitted_at': sub.submitted_at if sub else None,
                    'on_time': (sub.submitted_at.date() <= a.due_date) if sub else None,
                    'score': sub.score if sub else None,
                })
            category_breakdown.append({
                'category': cat_key,
                'label': cat_label,
                'total': assignments.count(),
                'submitted': sum(1 for r in rows if r['submitted']),
                'rows': rows,
            })

        eligibility = next((e for e in get_batch_eligibility(batch_id) if e['student_id'] == student.id), None)
        session = MockInterviewSession.objects.filter(batch=batch, student=student).first()

        # Pull this student's personal note out of any generated Report, most recent first
        student_notes = []
        for report in Report.objects.filter(batch=batch).order_by('-created_at'):
            try:
                content = json.loads(report.content)
            except (ValueError, TypeError):
                continue
            for note in content.get('student_notes', []):
                if note.get('student') == student.username:
                    student_notes.append({
                        'report_title': report.title,
                        'created_at': report.created_at,
                        'note': note.get('note'),
                    })

        return Response({
            'student': {
                'id': student.id, 'username': student.username, 'email': student.email,
                'first_name': student.first_name, 'last_name': student.last_name,
                'profile_photo': student.profile_photo, 'phone': student.phone,
                'personal_email': student.personal_email, 'official_email': student.official_email,
                'tenth_marksheet_url': student.tenth_marksheet.url if student.tenth_marksheet else None,
                'twelfth_marksheet_url': student.twelfth_marksheet.url if student.twelfth_marksheet else None,
                'degree_certificate_url': student.degree_certificate.url if student.degree_certificate else None,
                'pg_certificate_url': student.pg_certificate.url if student.pg_certificate else None,
                'terms_conditions_doc_url': student.terms_conditions_doc.url if student.terms_conditions_doc else None,
                
                'education': {
                    'tenth': {'school': student.tenth_school, 'year': student.tenth_year, 'percentage': student.tenth_percentage},
                    'twelfth': {'school': student.twelfth_school, 'year': student.twelfth_year, 'percentage': student.twelfth_percentage},
                    'ug': {'degree': student.ug_degree, 'college': student.ug_college, 'year': student.ug_year, 'percentage': student.ug_percentage},
                    'pg': {'degree': student.pg_degree, 'college': student.pg_college, 'year': student.pg_year, 'percentage': student.pg_percentage},
                },
            },
            'batch': {
                'id': batch.id, 'name': batch.name,
                'trainer_username': batch.trainer.username,
                'trainer_first_name': batch.trainer.first_name,
                'trainer_last_name': batch.trainer.last_name,
            },
            'attendance_percentage': attendance_pct,
            'present_days': present_days,
            'total_days': total_days,
            'attendance_records': [
                {'date': a.date, 'status': a.status} for a in attendance_qs.order_by('-date')[:30]
            ],
            'category_breakdown': category_breakdown,
            'eligible': eligibility['eligible'] if eligibility else False,
            'mock_interview': {
                'invited': session is not None,
                'scheduled_datetime': session.scheduled_datetime if session else None,
                'meeting_link': getattr(session, 'meeting_link', None) if session else None,
                'attended': session.attended if session else None,
                'score': session.score if session else None,
                'feedback': session.feedback if session else None,
            },
            'student_notes': student_notes,
        })

import openpyxl


class UploadMockScoresView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        batch_id = request.data.get('batch_id')
        file = request.FILES.get('file')

        if not batch_id or not file:
            return Response({"detail": "batch_id and file are required."}, status=400)

        try:
            workbook = openpyxl.load_workbook(file, data_only=True)
            sheet = workbook.active
        except Exception as e:
            return Response({"detail": f"Could not read Excel file: {e}"}, status=400)

        # Expect columns: Student (username), Attended (yes/no), Score
        header_row = [cell.value for cell in sheet[1]]
        try:
            student_col = header_row.index('Student')
            attended_col = header_row.index('Attended')
            score_col = header_row.index('Score')
        except ValueError:
            return Response(
                {"detail": "Excel must have columns named 'Student', 'Attended', and 'Score'."},
                status=400
            )

        updated = []
        errors = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            username = row[student_col]
            attended_raw = row[attended_col]
            score_raw = row[score_col]

            if not username:
                continue

            try:
                student = User.objects.get(username=str(username).strip(), role='student')
            except User.DoesNotExist:
                errors.append(f"No student found with username '{username}'.")
                continue

            session = MockInterviewSession.objects.filter(batch_id=batch_id, student=student).first()
            if not session:
                errors.append(f"No mock interview session found for '{username}' in this batch.")
                continue

            attended_value = str(attended_raw).strip().lower() in ('yes', 'true', '1')
            score_value = None
            if score_raw is not None and str(score_raw).strip() != '':
                try:
                    score_value = int(float(score_raw))
                except ValueError:
                    errors.append(f"Invalid score for '{username}': {score_raw}")
                    continue

            session.attended = attended_value
            session.score = score_value
            session.save()
            updated.append(username)

        return Response({
            "updated": updated,
            "updated_count": len(updated),
            "errors": errors,
        })  

class EnrollStudentView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        batch_id = request.data.get('batch_id')
        student_id = request.data.get('student_id')

        if not batch_id or not student_id:
            return Response({"detail": "batch_id and student_id are required."}, status=400)

        try:
            batch = Batch.objects.get(id=batch_id)
            student = User.objects.get(id=student_id, role='student')
        except (Batch.DoesNotExist, User.DoesNotExist):
            return Response({"detail": "Batch or student not found."}, status=404)

        # Check capacity
        current_count = (
            Attendance.objects.filter(batch=batch)
            .order_by('student_id')
            .values_list('student_id', flat=True)
            .distinct()
            .count()
        )
        already_enrolled = Attendance.objects.filter(batch=batch, student=student).exists()

        if not already_enrolled and current_count >= batch.max_students:
            return Response(
                {"detail": f"This batch has reached its maximum capacity of {batch.max_students} students."},
                status=400
            )

        if already_enrolled:
            return Response({"detail": f"{student.username} is already enrolled in this batch."}, status=200)

        # Create a placeholder attendance record marking enrollment (today, present by default)
        Attendance.objects.create(
            batch=batch,
            student=student,
            date=timezone.now().date(),
            status='present',
            marked_by=request.user,
        )

        return Response({"detail": f"{student.username} enrolled in {batch.name}."}, status=201)     


class BulkEnrollStudentsView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        batch_id = request.data.get('batch_id')
        file = request.FILES.get('file')

        if not batch_id or not file:
            return Response({"detail": "batch_id and file are required."}, status=400)

        try:
            batch = Batch.objects.get(id=batch_id)
        except Batch.DoesNotExist:
            return Response({"detail": "Batch not found."}, status=404)

        try:
            workbook = openpyxl.load_workbook(file, data_only=True)
            sheet = workbook.active
        except Exception as e:
            return Response({"detail": f"Could not read Excel file: {e}"}, status=400)

        header_row = [cell.value for cell in sheet[1]]
        try:
            username_col = header_row.index('Username')
            email_col = header_row.index('Email')
            password_col = header_row.index('Password')
        except ValueError:
            return Response(
                {"detail": "Excel must have columns named 'Username', 'Email', and 'Password'."},
                status=400
            )

        created = []
        enrolled = []
        errors = []

        current_count = (
            Attendance.objects.filter(batch=batch)
            .order_by('student_id')
            .values_list('student_id', flat=True)
            .distinct()
            .count()
        )

        for row in sheet.iter_rows(min_row=2, values_only=True):
            username = row[username_col]
            email = row[email_col]
            password = row[password_col]

            if not username or not email or not password:
                continue

            username = str(username).strip()
            email = str(email).strip()
            password = str(password).strip()

            student, was_created = User.objects.get_or_create(
                username=username,
                defaults={'email': email, 'role': 'student'},
            )
            if was_created:
                student.set_password(password)
                student.email = email
                student.role = 'student'
                student.save()
                created.append(username)
            else:
                if student.role != 'student':
                    errors.append(f"'{username}' already exists with a different role. Skipped.")
                    continue

            already_enrolled = Attendance.objects.filter(batch=batch, student=student).exists()
            if already_enrolled:
                errors.append(f"'{username}' is already enrolled in this batch.")
                continue

            if current_count >= batch.max_students:
                errors.append(f"Batch full. Could not enroll '{username}'.")
                continue

            Attendance.objects.create(
                batch=batch,
                student=student,
                date=timezone.now().date(),
                status='present',
                marked_by=request.user,
            )
            enrolled.append(username)
            current_count += 1

        return Response({
            "created": created,
            "enrolled": enrolled,
            "enrolled_count": len(enrolled),
            "errors": errors,
        })   


class MessageViewSet(viewsets.ModelViewSet):
    serializer_class = MessageSerializer
    permission_classes = [permissions.IsAuthenticated]

    def get_queryset(self):
        user = self.request.user
        if user.role in ('admin', 'management'):
            qs = Message.objects.all()
        else:
            qs = Message.objects.filter(models.Q(sender=user) | models.Q(recipient=user))
        student_id = self.request.query_params.get('student_id')
        if student_id:
            qs = qs.filter(models.Q(sender_id=student_id) | models.Q(recipient_id=student_id))
        return qs

    def perform_create(self, serializer):
        serializer.save(sender=self.request.user)

class UnreadMessageCountView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        count = Message.objects.filter(recipient=request.user, is_read=False).count()
        return Response({"unread_count": count})


class MarkMessagesReadView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        student_id = request.data.get('student_id')
        if not student_id:
            return Response({"detail": "student_id is required."}, status=400)
        Message.objects.filter(
            recipient=request.user, sender_id=student_id, is_read=False
        ).update(is_read=True)
        return Response({"detail": "Marked as read."})

class SendWelcomeEmailView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        if request.user.role not in ('admin', 'management'):
            return Response({"detail": "Not authorized."}, status=403)

        student_ids = request.data.get('student_ids', [])
        subject = (request.data.get('subject') or '').strip()
        body_template = request.data.get('body', '')
        cc_raw = request.data.get('cc', '')
        batch_id = request.data.get('batch_id')

        if not student_ids:
            return Response({"detail": "student_ids is required."}, status=400)
        if not subject or not body_template:
            return Response({"detail": "subject and body are required."}, status=400)

        cc_list = [c.strip() for c in cc_raw.split(',') if c.strip()] if cc_raw else None

        sent = []
        skipped = []

        for sid in student_ids:
            try:
                student = User.objects.get(id=sid, role='student')
            except User.DoesNotExist:
                skipped.append({"id": sid, "reason": "Student not found."})
                continue

            full_name = display_name(student)
            personalized_body = body_template.replace('{{full_name}}', full_name)

            recipients = [e for e in [student.personal_email, student.official_email] if e]
            if not recipients:
                skipped.append({"id": sid, "reason": f"No email on file for {full_name}."})
                continue

            if send_email(to=recipients, subject=subject, html_body=personalized_body, cc=cc_list):
                sent.append({"id": sid, "name": full_name})
            else:
                skipped.append({"id": sid, "reason": f"Email send failed for {full_name}."})

        if batch_id and len(sent) > 0:
            from trainer_ai.models import Batch
            Batch.objects.filter(id=batch_id).update(welcome_email_sent=True)

        return Response({"sent": sent, "sent_count": len(sent), "skipped": skipped})  


class NotifyTrainerView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        if request.user.role not in ('admin', 'management'):
            return Response({"detail": "Not authorized."}, status=403)

        to = (request.data.get('to') or '').strip()
        subject = (request.data.get('subject') or '').strip()
        body = request.data.get('body', '')
        cc_raw = request.data.get('cc', '')
        batch_id = request.data.get('batch_id')

        if not to or not subject or not body:
            return Response({"detail": "to, subject, and body are required."}, status=400)

        cc_list = [c.strip() for c in cc_raw.split(',') if c.strip()] if cc_raw else None

        if send_email(to=to, subject=subject, html_body=body, cc=cc_list):
            if batch_id:
                Batch.objects.filter(id=batch_id).update(trainer_notified=True)
            return Response({"detail": "Trainer notified."})
        return Response({"detail": "Failed to send email."}, status=500)  


# ---------------------------------------------------------------------------
# Absent Student Notifier
# ---------------------------------------------------------------------------

class AbsentStudentsView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        batch_id = request.query_params.get('batch_id')
        date = request.query_params.get('date')
        if not batch_id or not date:
            return Response({"detail": "batch_id and date are required."}, status=400)

        notified_ids = set(
            AbsenceNotification.objects.filter(batch_id=batch_id, date=date).values_list('student_id', flat=True)
        )

        absent_records = Attendance.objects.filter(batch_id=batch_id, date=date, status='absent')
        results = []
        for a in absent_records:
            s = a.student
            results.append({
                'id': s.id,
                'name': display_name(s),
                'personal_email': s.personal_email,
                'official_email': s.official_email,
                'already_notified': s.id in notified_ids,
            })
        return Response(results)


class NotifyAbsentStudentsView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        batch_id = request.data.get('batch_id')
        date = request.data.get('date')
        student_ids = request.data.get('student_ids', [])
        subject = (request.data.get('subject') or '').strip()
        body_template = request.data.get('body', '')
        cc_raw = request.data.get('cc', '')

        if not batch_id or not date:
            return Response({"detail": "batch_id and date are required."}, status=400)
        if not student_ids or not subject or not body_template:
            return Response({"detail": "student_ids, subject, and body are required."}, status=400)

        cc_list = [c.strip() for c in cc_raw.split(',') if c.strip()] if cc_raw else None

        sent = []
        skipped = []
        for sid in student_ids:
            try:
                student = User.objects.get(id=sid, role='student')
            except User.DoesNotExist:
                skipped.append({"id": sid, "reason": "Student not found."})
                continue

            full_name = display_name(student)
            personalized_body = body_template.replace('{{full_name}}', full_name)
            personalized_subject = subject.replace('{{full_name}}', full_name)

            to_email = student.personal_email or student.official_email
            if not to_email:
                skipped.append({"id": sid, "reason": f"No email on file for {full_name}."})
                continue

            if send_email(to=to_email, subject=personalized_subject, html_body=personalized_body, cc=cc_list):
                AbsenceNotification.objects.get_or_create(batch_id=batch_id, student_id=sid, date=date)
                sent.append({"id": sid, "name": full_name})
            else:
                skipped.append({"id": sid, "reason": f"Email send failed for {full_name}."})

        return Response({"sent": sent, "sent_count": len(sent), "skipped": skipped})


# ---------------------------------------------------------------------------
# Class Recordings
# ---------------------------------------------------------------------------

class ClassRecordingListCreateView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        batch_id = request.query_params.get('batch_id')
        if not batch_id:
            return Response({"detail": "batch_id is required."}, status=400)

        recordings = ClassRecording.objects.filter(batch_id=batch_id)
        data = []
        for r in recordings:
            views_qs = r.views.all()
            data.append({
                'id': r.id,
                'title': r.title,
                'date': r.date,
                'link': r.link,
                'notes': r.notes,
                'created_at': r.created_at,
                'sent_count': views_qs.count(),
                'watched_count': views_qs.filter(clicked=True).count(),
            })
        return Response(data)

    def post(self, request):
        batch_id = request.data.get('batch_id')
        date = request.data.get('date')
        title = request.data.get('title')
        link = request.data.get('link')
        notes = request.data.get('notes', '')

        if not all([batch_id, date, title, link]):
            return Response({"detail": "batch_id, date, title, and link are required."}, status=400)

        recording = ClassRecording.objects.create(
            batch_id=batch_id, date=date, title=title, link=link, notes=notes, created_by=request.user,
        )
        return Response({
            "id": recording.id, "title": recording.title, "date": recording.date,
            "link": recording.link, "notes": recording.notes,
        }, status=201)


class ShareRecordingView(APIView):
    """Sends the recording link to selected students' personal emails,
    routed through a per-student tracking redirect so we know who actually
    clicked it."""
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request, recording_id):
        try:
            recording = ClassRecording.objects.get(id=recording_id)
        except ClassRecording.DoesNotExist:
            return Response({"detail": "Recording not found."}, status=404)

        student_ids = request.data.get('student_ids', [])
        subject = (request.data.get('subject') or f"Class Recording: {recording.title}").strip()
        body_template = request.data.get('body', '')
        cc_raw = request.data.get('cc', '')

        if not student_ids:
            return Response({"detail": "student_ids is required."}, status=400)

        cc_list = [c.strip() for c in cc_raw.split(',') if c.strip()] if cc_raw else None

        base_url = request.build_absolute_uri('/').rstrip('/')
        sent = []
        skipped = []

        for sid in student_ids:
            try:
                student = User.objects.get(id=sid, role='student')
            except User.DoesNotExist:
                skipped.append({"id": sid, "reason": "Student not found."})
                continue

            to_email = student.personal_email or student.official_email
            if not to_email:
                skipped.append({"id": sid, "reason": "No email on file."})
                continue

            view_record, _ = RecordingView.objects.get_or_create(
                recording=recording, student=student,
                defaults={'token': secrets.token_urlsafe(24)},
            )
            tracked_link = f"{base_url}/api/trainer/recordings/track/{view_record.token}/"

            full_name = display_name(student)
            personalized_body = (body_template or f"<p>Hi {{full_name}},</p><p>Here's the recording for {recording.title}.</p>") \
                .replace('{{full_name}}', full_name) \
                .replace('{{recording_link}}', tracked_link)

            if '{{recording_link}}' not in body_template:
                personalized_body += f"<p><a href='{tracked_link}'>Watch Recording</a></p>"

            if send_email(to=to_email, subject=subject, html_body=personalized_body, cc=cc_list):
                sent.append({"id": sid, "name": full_name})
            else:
                skipped.append({"id": sid, "reason": "Email send failed."})

        return Response({"sent": sent, "sent_count": len(sent), "skipped": skipped})


class RecordingClickTrackView(APIView):
    """Public — no auth. This is the URL embedded in the email itself."""
    permission_classes = [permissions.AllowAny]

    def get(self, request, token):
        view_record = get_object_or_404(RecordingView, token=token)
        if not view_record.clicked:
            view_record.clicked = True
            view_record.clicked_at = timezone.now()
            view_record.save(update_fields=['clicked', 'clicked_at'])
        return HttpResponseRedirect(view_record.recording.link)


class RecordingStatsView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, recording_id):
        try:
            recording = ClassRecording.objects.get(id=recording_id)
        except ClassRecording.DoesNotExist:
            return Response({"detail": "Recording not found."}, status=404)

        batch_id = recording.batch_id

        student_ids = (
            Attendance.objects.filter(batch_id=batch_id)
            .order_by('student_id')
            .values_list('student_id', flat=True)
            .distinct()
        )

        attendance_map = {
            a.student_id: a.status
            for a in Attendance.objects.filter(batch_id=batch_id, date=recording.date)
        }
        view_map = {v.student_id: v for v in recording.views.all()}

        rows = []
        for sid in student_ids:
            student = User.objects.get(id=sid)
            att_status = attendance_map.get(sid, 'not_marked')
            view = view_map.get(sid)
            watched = bool(view and view.clicked)
            attended_live = att_status == 'present'

            rows.append({
                'student_id': sid,
                'name': display_name(student),
                'attendance_status': att_status,
                'attended_live': attended_live,
                'sent_recording': view is not None,
                'watched_recording': watched,
                'watched_at': view.clicked_at if view and view.clicked else None,
                'watched_without_attending': watched and not attended_live,
            })

        return Response({
            'recording_title': recording.title,
            'recording_date': recording.date,
            'sent_count': recording.views.count(),
            'watched_count': sum(1 for r in rows if r['watched_recording']),
            'attended_live_count': sum(1 for r in rows if r['attended_live']),
            'watched_without_attending_count': sum(1 for r in rows if r['watched_without_attending']),
            'students': rows,
        })


# ---------------------------------------------------------------------------
# Discontinuation / Dropout Tracking
# ---------------------------------------------------------------------------

class BatchEnrollmentStatusView(APIView):
    """For a batch, lists every enrolled student with their current status,
    absence streak, and whether they're a discontinuation candidate."""
    permission_classes = [permissions.IsAuthenticated]

    STREAK_THRESHOLD = 5  # consecutive absences to flag as a candidate

    def get(self, request, batch_id):
        student_ids = (
            Attendance.objects.filter(batch_id=batch_id)
            .order_by('student_id')
            .values_list('student_id', flat=True)
            .distinct()
        )

        results = []
        for sid in student_ids:
            student = User.objects.get(id=sid)
            enrollment, _ = Enrollment.objects.get_or_create(batch_id=batch_id, student_id=sid)

            records = Attendance.objects.filter(batch_id=batch_id, student_id=sid).order_by('-date')
            streak = 0
            for r in records:
                if r.status == 'absent':
                    streak += 1
                else:
                    break

            total_absent_days = records.filter(status='absent').count()

            results.append({
                'student_id': sid,
                'name': display_name(student),
                'status': enrollment.status,
                'discontinued_date': enrollment.discontinued_date,
                'discontinued_reason': enrollment.discontinued_reason,
                'current_absence_streak': streak,
                'total_absent_days': total_absent_days,
                'is_candidate': streak >= self.STREAK_THRESHOLD and enrollment.status == 'active',
            })
        return Response(results)


class MarkDiscontinuedView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        batch_id = request.data.get('batch_id')
        student_id = request.data.get('student_id')
        discontinued_date = request.data.get('discontinued_date')
        reason = request.data.get('reason', '')

        if not batch_id or not student_id:
            return Response({"detail": "batch_id and student_id are required."}, status=400)

        enrollment, _ = Enrollment.objects.get_or_create(batch_id=batch_id, student_id=student_id)
        enrollment.status = 'discontinued'
        enrollment.discontinued_date = discontinued_date or timezone.now().date()
        enrollment.discontinued_reason = reason
        enrollment.save()

        return Response({"detail": "Student marked as discontinued."})


class ReactivateStudentView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        batch_id = request.data.get('batch_id')
        student_id = request.data.get('student_id')

        if not batch_id or not student_id:
            return Response({"detail": "batch_id and student_id are required."}, status=400)

        enrollment, _ = Enrollment.objects.get_or_create(batch_id=batch_id, student_id=student_id)
        enrollment.status = 'active'
        enrollment.discontinued_date = None
        enrollment.discontinued_reason = None
        enrollment.save()

        return Response({"detail": "Student reactivated."})    


def get_week_range():
    today = date.today()
    monday = today - timedelta(days=today.weekday())
    sunday = monday + timedelta(days=6)
    return monday, sunday


def get_month_range():
    today = date.today()
    first = today.replace(day=1)
    if today.month == 12:
        next_month_first = today.replace(year=today.year + 1, month=1, day=1)
    else:
        next_month_first = today.replace(month=today.month + 1, day=1)
    last = next_month_first - timedelta(days=1)
    return first, last


def compute_zone_rows(batch_id, start, end):
    batch = Batch.objects.get(id=batch_id)
    eligibility_data = get_batch_eligibility(batch_id)

    total_daily_tasks = Assignment.objects.filter(batch_id=batch_id, category='task').count()
    total_mini_projects = Assignment.objects.filter(batch_id=batch_id, category='mini_project').count()
    total_main_projects = Assignment.objects.filter(batch_id=batch_id, category='main_project').count()
    total_seminars = Assignment.objects.filter(batch_id=batch_id, category='seminar').count()

    schedule = DailySchedule.objects.filter(batch_id=batch_id).first()
    timing = f"{schedule.start_time.strftime('%I:%M %p')} - {schedule.end_time.strftime('%I:%M %p')}" if schedule else "—"

    rows = []
    sno = 1
    for e in eligibility_data:
        sid = e['student_id']
        att_records_period = Attendance.objects.filter(batch_id=batch_id, student_id=sid, date__gte=start, date__lte=end)
        total_days_period = att_records_period.count()
        present_days_period = att_records_period.filter(status='present').count()

        completed_daily_tasks = AssignmentSubmission.objects.filter(
            student_id=sid, assignment__batch_id=batch_id, assignment__category='task',
            submitted_at__date__gte=start, submitted_at__date__lte=end,
        ).count()
        daily_task_pct = round((completed_daily_tasks / total_daily_tasks) * 100) if total_daily_tasks > 0 else 0

        completed_mini_projects = AssignmentSubmission.objects.filter(
            student_id=sid, assignment__batch_id=batch_id, assignment__category='mini_project',
            submitted_at__date__gte=start, submitted_at__date__lte=end,
        ).count()
        mini_project_pct = round((completed_mini_projects / total_mini_projects) * 100) if total_mini_projects > 0 else 0

        completed_main_projects = AssignmentSubmission.objects.filter(
            student_id=sid, assignment__batch_id=batch_id, assignment__category='main_project',
            submitted_at__date__gte=start, submitted_at__date__lte=end,
        ).count()
        main_project_pct = round((completed_main_projects / total_main_projects) * 100) if total_main_projects > 0 else 0

        completed_seminars = AssignmentSubmission.objects.filter(
            student_id=sid, assignment__batch_id=batch_id, assignment__category='seminar',
            submitted_at__date__gte=start, submitted_at__date__lte=end,
        ).count()
        seminar_pct = round((completed_seminars / total_seminars) * 100) if total_seminars > 0 else 0

        zone = 'Safe Zone' if e['eligible'] else 'Danger Zone'
        trainee_full_name = f"{e['first_name']} {e['last_name']}".strip() or e['username']
        rows.append({
            'sno': sno, 'trainer_name': display_name(batch.trainer), 'trainee_name': trainee_full_name,
            'zone': zone, 'batch': batch.course_name or batch.name, 'timings': timing,
            'total_class_days': total_days_period, 'total_present_days': present_days_period,
            'attendance_percentage': e['attendance_percentage'],
            'assigned_daily_tasks': total_daily_tasks, 'completed_daily_tasks': completed_daily_tasks,
            'daily_task_percentage': daily_task_pct,
            'assigned_mini_projects': total_mini_projects, 'completed_mini_projects': completed_mini_projects,
            'mini_project_percentage': mini_project_pct,
            'assigned_main_projects': total_main_projects, 'completed_main_projects': completed_main_projects,
            'main_project_percentage': main_project_pct,
            'assigned_seminars': total_seminars, 'completed_seminars': completed_seminars,
            'seminar_percentage': seminar_pct,
        })
        sno += 1
    return rows


class EmailZoneReportView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def post(self, request):
        batch_id = request.data.get('batch_id')
        period = request.data.get('period')

        if not batch_id or period not in ('weekly', 'monthly'):
            return Response({"detail": "batch_id and a valid period ('weekly' or 'monthly') are required."}, status=400)

        batch = Batch.objects.get(id=batch_id)
        start, end = get_week_range() if period == 'weekly' else get_month_range()

        rows = compute_zone_rows(batch_id, start, end)
        if not rows:
            return Response({"detail": "No data available to report for this period."}, status=400)

        label = 'Weekly' if period == 'weekly' else 'Monthly'
        title = f"{batch.course_name or batch.name} - {label} Production Report ({start} to {end})"
        excel_buffer = build_zone_report_excel(rows, title)
        filename = f"{batch.name.replace(' ', '_')}_{label}_Zone_Report.xlsx"
        mimetype = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

        student_ids = (
            Attendance.objects.filter(batch_id=batch_id)
            .order_by('student_id').values_list('student_id', flat=True).distinct()
        )

        sent = []
        skipped = []
        for sid in student_ids:
            student = User.objects.get(id=sid)
            to_email = student.personal_email or student.official_email
            if not to_email:
                skipped.append(display_name(student))
                continue
            excel_buffer.seek(0)
            success = send_email(
                to=to_email,
                subject=f"{label} Report - {batch.course_name or batch.name}",
                html_body=f"<p>Hi {display_name(student)},</p><p>Please find attached your batch's {label.lower()} performance report ({start} to {end}).</p>",
                attachments=[(filename, excel_buffer.read(), mimetype)],
            )
            (sent if success else skipped).append(display_name(student))

        return Response({"sent_count": len(sent), "skipped_count": len(skipped), "skipped": skipped})    
    

class HolidayViewSet(viewsets.ModelViewSet):
    serializer_class = HolidaySerializer
    permission_classes = [permissions.IsAuthenticated]
    queryset = Holiday.objects.all()

    def get_queryset(self):
        batch_id = self.request.query_params.get('batch_id')
        qs = Holiday.objects.all()
        if batch_id:
            qs = qs.filter(batch_id=batch_id)
        return qs    