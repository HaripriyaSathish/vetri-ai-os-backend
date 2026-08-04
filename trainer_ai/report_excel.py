from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

TABLE_HEADERS = [
    'S.No', 'Trainer Name', 'Trainee Name', 'Zone', 'Batch', 'Training Taken Timings',
    'Total No. of Class Days', 'Total Present Days', 'Total Attendance %',
    'Assigned Daily Tasks', 'Total Daily Tasks Completed', 'Daily Tasks Completed %',
    'Assigned Mini Projects', 'Total Mini Project Completed', 'Mini Project Completed %',
]

COL_WIDTHS = [6, 14, 18, 12, 14, 18, 14, 14, 14, 16, 18, 16, 16, 18, 16]
ZONE_HIGHLIGHT_COLS = [4, 9, 12, 15]  # 1-indexed: Zone, Attendance %, Daily Tasks %, Mini Project %


def build_zone_report_excel(rows, title):
    wb = Workbook()
    ws = wb.active
    ws.title = 'Report'
    n_cols = len(TABLE_HEADERS)

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.fill = PatternFill('solid', fgColor='B8CCE4')
    title_cell.font = Font(bold=True, size=14)
    title_cell.alignment = Alignment(horizontal='center', vertical='center')

    thin = Side(style='thin', color='000000')
    border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for col_idx, header in enumerate(TABLE_HEADERS, start=1):
        cell = ws.cell(row=2, column=col_idx, value=header)
        cell.fill = PatternFill('solid', fgColor='FFFF99')
        cell.font = Font(bold=True, size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = border

    for row_idx, r in enumerate(rows, start=3):
        is_safe = r['zone'] == 'Safe Zone'
        row_color = 'C6EFCE' if is_safe else 'FFC7CE'
        font_color = '006100' if is_safe else '9C0006'

        values = [
            r['sno'], r['trainer_name'], r['trainee_name'], r['zone'], r['batch'], r['timings'],
            r['total_class_days'], r['total_present_days'], f"{r['attendance_percentage']}%",
            r['assigned_daily_tasks'], r['completed_daily_tasks'], f"{r['daily_task_percentage']}%",
            r['assigned_mini_projects'], r['completed_mini_projects'], f"{r['mini_project_percentage']}%",
        ]
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if col_idx in ZONE_HIGHLIGHT_COLS:
                cell.fill = PatternFill('solid', fgColor=row_color)
                cell.font = Font(bold=True, color=font_color)
                cell.alignment = Alignment(horizontal='center')

    for col_idx, width in enumerate(COL_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer